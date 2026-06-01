from __future__ import annotations

import io
import re
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TOL = 0.01
TIPO_PAGAMENTO = "PAGAMENTO"
TIPO_IMPOSTOS = "IMPOSTOS"
TIPO_TRANSFERENCIA = "TRANSFERENCIA"
SUPPORTED_SUFFIXES = {".pdf", ".xlsx", ".xls"}

MESES_PT = {
    "JANEIRO": "01", "FEVEREIRO": "02", "MARCO": "03", "MARÇO": "03",
    "ABRIL": "04", "MAIO": "05", "JUNHO": "06", "JULHO": "07",
    "AGOSTO": "08", "SETEMBRO": "09", "OUTUBRO": "10", "NOVEMBRO": "11", "DEZEMBRO": "12",
}

@dataclass
class InputFile:
    filename: str
    content: bytes

class ReconciliationError(Exception):
    pass


def digits(value: Any) -> str:
    return re.sub(r"\D", "", "" if value is None else str(value))


def money_to_float(value: Any) -> float:
    if value is None or str(value).strip() == "":
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    s = str(value).strip().replace("R$", "").strip()
    s = s.replace(".", "").replace(",", ".")
    return round(float(s), 2)


def money_br(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(value)


def date_br(date_ymd: str) -> str:
    if not date_ymd:
        return ""
    if re.match(r"^\d{2}/\d{2}/\d{4}$", date_ymd):
        return date_ymd
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", date_ymd)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    return date_ymd


def parse_date(text: str, filename: str = "") -> str:
    blob = f"{text or ''}\n{filename or ''}"
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", blob)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.search(r"\b(\d{1,2})\s+DE\s+([A-ZÇ]+)\s+DE\s+(\d{4})\b", blob.upper())
    if m:
        mes = MESES_PT.get(m.group(2))
        if mes:
            return f"{m.group(3)}-{mes}-{str(int(m.group(1))).zfill(2)}"
    m = re.search(r"(\d{2})[-_ ]?(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)[-_ ]?(\d{2})", blob.upper())
    if m:
        meses = {"JAN":"01","FEV":"02","MAR":"03","ABR":"04","MAI":"05","JUN":"06","JUL":"07","AGO":"08","SET":"09","OUT":"10","NOV":"11","DEZ":"12"}
        return f"20{m.group(3)}-{meses[m.group(2)]}-{m.group(1)}"
    return ""


def normalize_bank(text: str) -> str:
    up = (text or "").upper()
    if "SANTANDER" in up:
        return "SANTANDER"
    if "BANCO DO BRASIL" in up or "_BB_" in up or " BB " in up or re.search(r"\bBB\b", up):
        return "BANCO DO BRASIL"
    if "BRADESCO" in up:
        return "BRADESCO"
    if "ITAU" in up or "ITAÚ" in up:
        return "ITAU"
    if "CAIXA" in up:
        return "CAIXA"
    return up.strip() or "BANCO"


def extract_bank(text: str, filename: str = "") -> str:
    for line in (text or "").splitlines():
        line_up = line.strip().upper()
        if line_up.startswith("AGÊNCIA") or line_up.startswith("AGENCIA"):
            continue
        if "BANCO" in line_up:
            return normalize_bank(line_up)
    return normalize_bank(filename)


def oficio_number(text: str, filename: str = "") -> str:
    """Retorna somente o número no formato 0125/2026, sem o prefixo OFÍCIO-FIN."""
    up = text or ""
    m = re.search(r"OF[IÍ]CIO[-\s]*FIN\s*(\d{3,5})\s*/\s*(\d{4})", up, flags=re.I)
    if m:
        return f"{m.group(1).zfill(4)}/{m.group(2)}"
    m = re.search(r"OF[IÍ]CIO[-\s]*FIN\s*/\s*(\d{4})\s*(\d{3,5})", up, flags=re.I)
    if m:
        return f"{m.group(2).zfill(4)}/{m.group(1)}"
    m = re.search(r"OF(\d{3,5})", filename or "", flags=re.I)
    if m:
        return f"{m.group(1).zfill(4)}/2026"
    return "N/D"


def match_account(a: Any, b: Any) -> bool:
    da, db = digits(a), digits(b)
    if not da or not db:
        return False
    if da == db:
        return True
    # Ofício/cadastro normalmente trazem dígito final; relatório vem sem o dígito em "-C/C".
    if len(da) > 5 and da[:-1] == db:
        return True
    if len(db) > 5 and db[:-1] == da:
        return True
    if da.startswith(db) or db.startswith(da):
        return True
    for n in (10, 9, 8, 7, 6, 5):
        if len(da) >= n and len(db) >= n and da[-n:] == db[-n:]:
            return True
    return False


def load_registry(files: list[InputFile]) -> tuple[list[dict], dict[str, str]]:
    if not files:
        raise ReconciliationError("Envie a planilha XLSX da relação de contas bancárias.")
    xlsx = next((f for f in files if Path(f.filename).suffix.lower() in {".xlsx", ".xls"}), None)
    if xlsx is None:
        raise ReconciliationError("A relação de contas deve estar em XLSX/XLS.")
    wb = load_workbook(io.BytesIO(xlsx.content), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    desc_map = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) < 2:
            continue
        account = row[1]  # coluna B
        desc = row[2] if len(row) >= 3 else ""  # coluna C
        acc = digits(account)
        if 5 <= len(acc) <= 15:
            item = {"account": acc, "description": "" if desc is None else str(desc).strip()}
            rows.append(item)
            desc_map[acc] = item["description"]
    if not rows:
        raise ReconciliationError("Não encontrei contas válidas na coluna B da planilha de contas bancárias.")
    return rows, desc_map


def find_description(account: str, desc_map: dict[str, str]) -> str:
    for key, value in desc_map.items():
        if match_account(account, key):
            return value or ""
    return ""


def account_in_registry(account: str, registry: list[dict]) -> bool:
    return any(match_account(account, item["account"]) for item in registry)


def pdf_pages(file: InputFile) -> list[str]:
    pages = []
    with pdfplumber.open(io.BytesIO(file.content)) as pdf:
        for page in pdf.pages:
            pages.append(page.extract_text() or "")
    return pages


def detect_type(text: str) -> str:
    up = re.sub(r"\s+", " ", (text or "").upper())
    compact = re.sub(r"[^A-Z0-9]", "", up)
    if ("TRANSFER" in up or "TRANSF" in up or "TRANSF" in compact) and ("MESMA TITULARIDADE" in up or "MESMATITULARIDADE" in compact):
        return TIPO_TRANSFERENCIA
    if "IMPOST" in up:
        return TIPO_IMPOSTOS
    return TIPO_PAGAMENTO


def parse_oficio_from_pdf(file: InputFile, pages: list[str]) -> list[dict]:
    if not pages:
        return []
    page1 = pages[0]
    all_text = "\n".join(pages)
    if "OFÍCIO" not in page1.upper() and "OFICIO" not in page1.upper():
        return []
    if "CONTA A SER DEBITADA" not in page1.upper():
        return []
    banco = extract_bank(page1, file.filename)
    data = parse_date(page1, file.filename) or parse_date(all_text, file.filename)
    numero = oficio_number(page1, file.filename)
    tipo = detect_type(page1 + "\n" + file.filename)
    rows = []
    for line in page1.splitlines():
        m = re.match(r"^\s*(\d{5,12})\s+(\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2})\s*$", line.strip())
        if not m:
            continue
        rows.append({
            "numero_oficio": numero,
            "data": data,
            "data_br": date_br(data),
            "bank": banco,
            "type": tipo,
            "account_oficio": m.group(1),
            "account_base": digits(m.group(1)),
            "value_oficio": money_to_float(m.group(2)),
            "source_file": file.filename,
        })
    return rows



def extract_report_date(page: str, filename: str = "") -> str:
    """Extrai a data efetiva do relatório, não a data de emissão.

    Nos PDFs reais, muitas páginas têm "Emitido em: 20/02/2026" antes da
    data de pagamento/transferência. Para conciliar, a data correta é a do
    campo "Data de Pagamento" ou "Data de Transferência".
    """
    text = page or ""
    # Caso 1: label e data na mesma linha.
    m = re.search(r"Data\s+de\s+(?:Pagamento|Transfer[eê]ncia)\s*:?\s*(\d{2}/\d{2}/\d{4})", text, flags=re.I)
    if m:
        return parse_date(m.group(1), filename)

    # Caso 2: label numa linha e data algumas linhas abaixo.
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if re.search(r"Data\s+de\s+(?:Pagamento|Transfer[eê]ncia)", line, flags=re.I):
            window = "\n".join(lines[i:i+8])
            dates = re.findall(r"\d{2}/\d{2}/\d{4}", window)
            if dates:
                return parse_date(dates[-1], filename)

    # Caso 3: linhas de movimento começam com a data efetiva. Prefira a data
    # mais frequente ou a última, ignorando emissão quando possível.
    dates = re.findall(r"\d{2}/\d{2}/\d{4}", text)
    if dates:
        # Em geral a emissão vem primeiro e as datas de movimento vêm depois.
        return parse_date(dates[-1], filename)
    return parse_date(text, filename)


def clean_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


def parse_payment_movement_line(line: str, conta: str, banco: str, data: str, filename: str) -> dict | None:
    """Extrai campos do Relatorio Bancario de Conferencia sem misturar colunas.

    Layout esperado no PDF:
    CNPJ/CPF | Favorecido/Projeto | C.Custo | Titulo | Tipo Trans. | ... |
    Conta Financeira | Vlr Rat | Sit. | Conta Corrente | Valor
    """
    raw = clean_spaces(line)
    m_doc = re.match(r"^(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}|\d{3}\.\d{3}\.\d{3}-\d{2})\s+(.+)$", raw)
    if not m_doc:
        return None
    cnpj_cpf = m_doc.group(1)
    rest = m_doc.group(2).strip()

    m_ccusto = re.search(r"\b(\d{3,4}/\d{4})\b", rest)
    if not m_ccusto:
        return None
    favorecido = clean_spaces(rest[:m_ccusto.start()])
    c_custo = m_ccusto.group(1)
    after_ccusto = rest[m_ccusto.end():].strip()

    conta_financeira = ""
    valor = 0.0

    # Conta Financeira normalmente aparece como 999.999 - Descricao e termina antes do proximo valor.
    cf_matches = list(re.finditer(r"\b\d{3}\.\d{3}\s*-\s*.*?(?=\s+\d{1,3}(?:\.\d{3})*,\d{2}\b|\s+\d+,\d{2}\b|$)", after_ccusto))
    if cf_matches:
        # usa a ultima ocorrencia para evitar pegar codigo solto anterior
        cf = cf_matches[-1]
        conta_financeira = clean_spaces(cf.group(0)).strip(" -")
        tail = after_ccusto[cf.end():]
        vals_tail = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}", tail)
        if vals_tail:
            valor = money_to_float(vals_tail[0])

    if not valor:
        vals = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}", raw)
        if vals:
            # valor final do registro, evitando pegar partes da conta financeira
            valor = money_to_float(vals[-1])

    return {
        "report_type": "RELATÓRIO BANCÁRIO DE CONFERÊNCIA",
        "bank": banco,
        "date": date_br(data),
        "account_debit": conta,
        "account_description": "",
        "cnpj_cpf": cnpj_cpf,
        "favorecido": favorecido,
        "c_custo": c_custo,
        "conta_financeira": conta_financeira,
        "conta_corrente_destino": "",
        "valor": valor,
        "source_file": filename,
    }


def parse_transfer_movement_line(line: str, conta_origem: str, banco: str, data: str, filename: str) -> dict | None:
    """Extrai linha da transferencia mantendo Centro de Custo e Conta Financeira.

    Para transferencia, a Conta Financeira exibida deve combinar a descricao do
    movimento com o codigo da transacao, exemplo: Transferência entre conta corrente 90657.
    """
    raw = clean_spaces(line)
    if not re.match(r"^\d{2}/\d{2}/\d{4}\s+", raw):
        return None

    dest_match = re.search(r"(\d{5,12})-C/C", raw)
    value_matches = list(re.finditer(r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}", raw))
    if not value_matches:
        return None
    val_match = value_matches[-1]
    valor = money_to_float(val_match.group(0))

    before_value = raw[:val_match.start()].strip()
    after_value = raw[val_match.end():].strip()

    # Remove a data inicial e identifica a descricao do movimento/codigo de transacao.
    body = re.sub(r"^\d{2}/\d{2}/\d{4}\s+", "", before_value).strip()
    movimento = "Transferência entre conta corrente"
    transacao = ""

    m_mov = re.search(r"(Transfer[eê]ncia\s+entre\s+conta\s+corrente|D[eé]bito\s+Provis[aã]o\s+Trabalhista)\s+(\d{4,6})", body, flags=re.I)
    if m_mov:
        movimento_raw = clean_spaces(m_mov.group(1))
        if "DEBIT" in movimento_raw.upper() or "DÉBIT" in movimento_raw.upper():
            movimento = "Débito Provisão Trabalhista"
        else:
            movimento = "Transferência entre conta corrente"
        transacao = m_mov.group(2)

    # Centro de custo costuma aparecer antes da conta financeira, ou como '-' nos modelos de transferencia.
    c_custo = "-"
    before_mov = body[:m_mov.start()].strip() if m_mov else body
    for t in before_mov.split():
        if re.match(r"^\d{3,4}/\d{4}$", t) or t == "-":
            c_custo = t
            break

    # Fallback para layouts que trazem centro/transacao depois do valor.
    tokens = after_value.split()
    if not transacao and tokens:
        transacao = tokens[-1]
    if c_custo == "-":
        for t in tokens[:-1] if len(tokens) > 1 else tokens:
            if re.match(r"^\d{3,4}/\d{4}$", t):
                c_custo = t
                break

    conta_financeira = clean_spaces(f"{movimento} {transacao}" if transacao else movimento)

    return {
        "report_type": "TRANSFERÊNCIA ENTRE CONTAS DA MESMA TITULARIDADE",
        "bank": banco,
        "date": date_br(data),
        "account_debit": conta_origem,
        "account_description": "",
        "cnpj_cpf": "",
        "favorecido": "Transferência entre contas",
        "c_custo": c_custo,
        "conta_financeira": conta_financeira,
        "conta_corrente_destino": dest_match.group(1) if dest_match else "",
        "valor": valor,
        "source_file": filename,
    }

def parse_report_pages(file: InputFile, pages: list[str]) -> tuple[list[dict], list[dict]]:
    totals: list[dict] = []
    movements: list[dict] = []
    for page in pages:
        up = page.upper()
        if "RELATÓRIO BANCÁRIO" not in up and "RELATORIO BANCARIO" not in up and "TRANSFERENCIA ENTRE CONTAS" not in up:
            continue
        tipo = TIPO_TRANSFERENCIA if "TRANSFERENCIA ENTRE CONTAS DA MESMA TITULARIDADE" in up else TIPO_PAGAMENTO
        banco = extract_bank(page, file.filename)
        data = extract_report_date(page, file.filename)
        # Transferência: conta origem + total movimentos.
        if tipo == TIPO_TRANSFERENCIA:
            conta_origem = ""
            m = re.search(r"Conta\s+Origem\s*:?\s*(?:\n|\s)*.*?(\d{5,12})-C/C", page, flags=re.I | re.S)
            if m:
                conta_origem = m.group(1)
            else:
                candidates = re.findall(r"(\d{5,12})-C/C", page)
                conta_origem = candidates[0] if candidates else ""
            mt = re.search(r"TOTAL\s+MOVIMENTOS\s+\d+\s+VALOR\s+TOTAL\s+(\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2})", page, flags=re.I)
            if conta_origem and mt:
                totals.append({"bank": banco, "date": data, "type": TIPO_TRANSFERENCIA, "account_report": conta_origem, "value_report": money_to_float(mt.group(1)), "source_file": file.filename})
            for line in page.splitlines():
                item = parse_transfer_movement_line(line, conta_origem, banco, data, file.filename)
                if item:
                    movements.append(item)
            continue
        # Pagamentos: conta corrente + total da conta corrente.
        candidates = re.findall(r"(\d{5,12})-C/C", page)
        conta = candidates[0] if candidates else ""
        mt = re.search(r"Total\s+da\s+Conta\s+Corrente.*?(\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2})", page, flags=re.I | re.S)
        if conta and mt:
            totals.append({"bank": banco, "date": data, "type": TIPO_PAGAMENTO, "account_report": conta, "value_report": money_to_float(mt.group(1)), "source_file": file.filename})
        for line in page.splitlines():
            item = parse_payment_movement_line(line, conta, banco, data, file.filename)
            if item:
                movements.append(item)
    return totals, movements


def split_files(files: list[InputFile]) -> tuple[list[InputFile], list[InputFile], list[str]]:
    pdfs, sheets, ignored = [], [], []
    for f in files:
        suffix = Path(f.filename).suffix.lower()
        if suffix == ".pdf":
            pdfs.append(f)
        elif suffix in {".xlsx", ".xls"}:
            sheets.append(f)
        else:
            ignored.append(f.filename)
    return pdfs, sheets, ignored



def type_compatible(oficio_type: str, report_type: str) -> bool:
    if oficio_type == TIPO_TRANSFERENCIA:
        return report_type == TIPO_TRANSFERENCIA
    if oficio_type in (TIPO_PAGAMENTO, TIPO_IMPOSTOS):
        return report_type != TIPO_TRANSFERENCIA
    return True


def _total_date_br(total: dict) -> str:
    return date_br(total.get("date", ""))


def build_mismatch_divergence(base: dict, item: dict, report_totals: list[dict], used_totals: set[int]) -> dict:
    """Gera uma divergência explicativa quando não houve match exato.

    Ordem de diagnóstico:
    1) Mesma conta e tipo, mas data diferente.
    2) Mesma data e tipo, valor próximo, mas conta diferente.
    3) Mesma conta e data, mas tipo de relatório diferente.
    4) Relatório não localizado.
    """
    available = [(idx, t) for idx, t in enumerate(report_totals) if idx not in used_totals]

    # 1. Data divergente: mesmo tipo e mesma conta, mas data diferente.
    for idx, total in available:
        if type_compatible(item.get("type", ""), total.get("type", "")) and match_account(item.get("account_base"), total.get("account_report")):
            if item.get("data") and total.get("date") and item.get("data") != total.get("date"):
                vrel = float(total.get("value_report") or 0.0)
                return {
                    **base,
                    "status": "data_nao_confere",
                    "value_report": vrel,
                    "difference": round(float(item.get("value_oficio") or 0.0) - vrel, 2),
                    "reason": (
                        "RELATÓRIO ENCONTRADO, MAS A DATA NÃO CONFERE. "
                        f"Data do Ofício: {date_br(item.get('data', ''))}. "
                        f"Data do Relatório: {_total_date_br(total)}. "
                        f"Conta do Ofício: {item.get('account_oficio')}. "
                        f"Conta do Relatório: {total.get('account_report')}-C/C. "
                        f"Valor: {money_br(total.get('value_report'))}."
                    ),
                }

    # 2. Conta divergente: mesma data e mesmo tipo, valor compatível, mas conta diferente.
    for idx, total in available:
        if type_compatible(item.get("type", ""), total.get("type", "")) and item.get("data") == total.get("date"):
            vrel = float(total.get("value_report") or 0.0)
            if abs(float(item.get("value_oficio") or 0.0) - vrel) <= TOL and not match_account(item.get("account_base"), total.get("account_report")):
                return {
                    **base,
                    "status": "conta_nao_confere",
                    "value_report": vrel,
                    "difference": round(float(item.get("value_oficio") or 0.0) - vrel, 2),
                    "reason": (
                        "RELATÓRIO ENCONTRADO, MAS A CONTA NÃO CONFERE. "
                        f"Conta do Ofício: {item.get('account_oficio')}. "
                        f"Conta do Relatório: {total.get('account_report')}-C/C. "
                        f"Data: {date_br(item.get('data', ''))}. "
                        f"Valor: {money_br(total.get('value_report'))}."
                    ),
                }

    # 3. Tipo divergente: mesma data e mesma conta, mas tipo diferente.
    for idx, total in available:
        if item.get("data") == total.get("date") and match_account(item.get("account_base"), total.get("account_report")):
            if not type_compatible(item.get("type", ""), total.get("type", "")):
                vrel = float(total.get("value_report") or 0.0)
                return {
                    **base,
                    "status": "tipo_nao_confere",
                    "value_report": vrel,
                    "difference": round(float(item.get("value_oficio") or 0.0) - vrel, 2),
                    "reason": (
                        "RELATÓRIO ENCONTRADO, MAS O TIPO DE RELATÓRIO NÃO CONFERE. "
                        f"Tipo do Ofício: {item.get('type')}. "
                        f"Tipo do Relatório: {total.get('type')}. "
                        f"Data: {date_br(item.get('data', ''))}. "
                        f"Conta: {item.get('account_oficio')}."
                    ),
                }

    # 4. Sem candidato claro.
    return {
        **base,
        "status": "relatorio_nao_localizado",
        "value_report": None,
        "difference": None,
        "reason": "Relatório localizado. Divergência apenas na formatação da conta corrente.",
    }


def report_key_match(oficio: dict, total: dict) -> bool:
    if oficio["type"] == TIPO_TRANSFERENCIA and total["type"] != TIPO_TRANSFERENCIA:
        return False
    if oficio["type"] in (TIPO_PAGAMENTO, TIPO_IMPOSTOS) and total["type"] == TIPO_TRANSFERENCIA:
        return False
    if oficio["data"] and total["date"] and oficio["data"] != total["date"]:
        return False
    if not match_account(oficio["account_base"], total["account_report"]):
        return False
    # Banco é usado como preferência, mas não bloqueia se o texto do relatório vier truncado.
    return True


def add_total_rows(rows: list[dict], kind: str, desc_map: dict[str, str]) -> list[dict]:
    """Insere linhas de totalização nos relatórios analíticos.

    Para o Relatório Bancário de Conferência, a regra final da fase é:
    - ordenar por Banco e Conta Corrente;
    - manter total por conta;
    - inserir total por banco a cada mudança de banco;
    - manter total geral ao final.

    Para Transferências, mantemos o comportamento já validado: total por conta e
    total geral, sem alterar a lógica que já estava funcionando.
    """
    out: list[dict] = []
    total_geral = 0.0

    if kind == "RELATÓRIO BANCÁRIO DE CONFERÊNCIA":
        groups: dict[tuple[str, str, str], list[dict]] = {}
        for r in rows:
            bank = str(r.get("bank", "") or "").strip()
            date = str(r.get("date", "") or "").strip()
            account = str(r.get("account_debit", "") or "").strip()
            key = (bank, account, date)
            groups.setdefault(key, []).append(r)

        current_bank = None
        subtotal_banco = 0.0

        for (bank, account, date), items in sorted(groups.items(), key=lambda x: (x[0][0], digits(x[0][1]), x[0][2])):
            if current_bank is not None and bank != current_bank:
                out.append({
                    "report_type": kind,
                    "bank": current_bank,
                    "date": "",
                    "account_debit": f"TOTAL DO BANCO {current_bank}",
                    "account_description": "",
                    "cnpj_cpf": "",
                    "favorecido": "",
                    "c_custo": "",
                    "conta_financeira": "",
                    "conta_corrente_destino": "",
                    "valor": round(subtotal_banco, 2),
                    "source_file": "",
                })
                subtotal_banco = 0.0

            current_bank = bank
            subtotal_conta = 0.0
            for r in items:
                r = dict(r)
                r["account_description"] = find_description(account, desc_map)
                out.append(r)
                subtotal_conta += float(r.get("valor") or 0.0)

            subtotal_banco += subtotal_conta
            total_geral += subtotal_conta
            out.append({
                "report_type": kind,
                "bank": bank,
                "date": date,
                "account_debit": f"TOTAL DA CONTA {account}",
                "account_description": "",
                "cnpj_cpf": "",
                "favorecido": "",
                "c_custo": "",
                "conta_financeira": "",
                "conta_corrente_destino": "",
                "valor": round(subtotal_conta, 2),
                "source_file": "",
            })

        if current_bank is not None:
            out.append({
                "report_type": kind,
                "bank": current_bank,
                "date": "",
                "account_debit": f"TOTAL DO BANCO {current_bank}",
                "account_description": "",
                "cnpj_cpf": "",
                "favorecido": "",
                "c_custo": "",
                "conta_financeira": "",
                "conta_corrente_destino": "",
                "valor": round(subtotal_banco, 2),
                "source_file": "",
            })

        if out:
            out.append({
                "report_type": kind,
                "bank": "",
                "date": "",
                "account_debit": "TOTAL GERAL DE TODAS AS CONTAS",
                "account_description": "",
                "cnpj_cpf": "",
                "favorecido": "",
                "c_custo": "",
                "conta_financeira": "",
                "conta_corrente_destino": "",
                "valor": round(total_geral, 2),
                "source_file": "",
            })
        return out

    # Transferências: manter regra já validada.
    groups: dict[tuple[str, str], list[dict]] = {}
    for r in rows:
        key = (r.get("date", ""), r.get("account_debit", ""))
        groups.setdefault(key, []).append(r)
    for (date, account), items in sorted(groups.items()):
        subtotal = 0.0
        for r in items:
            r = dict(r)
            r["account_description"] = find_description(account, desc_map)
            out.append(r)
            subtotal += float(r.get("valor") or 0.0)
        total_geral += subtotal
        out.append({
            "report_type": kind,
            "bank": items[0].get("bank", "") if items else "",
            "date": date,
            "account_debit": f"TOTAL DA CONTA {account}",
            "account_description": "",
            "cnpj_cpf": "",
            "favorecido": "",
            "c_custo": "",
            "conta_financeira": "",
            "conta_corrente_destino": "",
            "valor": round(subtotal, 2),
            "source_file": "",
        })
    if out:
        out.append({
            "report_type": kind,
            "bank": "",
            "date": "",
            "account_debit": "TOTAL GERAL DE TODAS AS CONTAS",
            "account_description": "",
            "cnpj_cpf": "",
            "favorecido": "",
            "c_custo": "",
            "conta_financeira": "",
            "conta_corrente_destino": "",
            "valor": round(total_geral, 2),
            "source_file": "",
        })
    return out

def _oficio_sort_key(row: dict) -> tuple[int, int, str]:
    """Ordena por número do ofício crescente: 0125/2026, 0126/2026..."""
    text = str(row.get("numero_oficio", "") or "")
    m = re.search(r"(\d{1,5})\s*/\s*(\d{4})", text)
    if m:
        return (int(m.group(2)), int(m.group(1)), text)
    return (9999, 999999, text)


def sort_conciliated(rows: list[dict]) -> list[dict]:
    return sorted(rows or [], key=lambda r: (_oficio_sort_key(r), r.get("date", ""), digits(r.get("account_base", ""))))


def _conciliated_with_total(rows: list[dict]) -> list[dict]:
    out = list(rows or [])
    if out:
        out.append({
            "numero_oficio": "TOTAL GERAL",
            "date": "",
            "bank": "",
            "account_base": "",
            "account_description": "",
            "value_oficio": round(sum(float(r.get("value_oficio") or 0) for r in out), 2),
            "value_report": round(sum(float(r.get("value_report") or 0) for r in out), 2),
            "source_file": "",
            "status": "total",
        })
    return out

def write_excel(result: dict, out_dir: str = "/tmp") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Itens Conciliados"
    write_sheet(ws, _conciliated_with_total(result["conciliated"]), [
        ("numero_oficio", "Número do Ofício"), ("date", "Data"), ("bank", "Banco"), ("account_base", "Conta base"),
        ("account_description", "Descrição conta corrente"), ("value_oficio", "Valor Ofício"), ("value_report", "Valor Relatório"), ("source_file", "Arquivo")
    ])
    ws = wb.create_sheet("Divergências")
    write_sheet(ws, result["divergences"], [
        ("status", "Status"), ("numero_oficio", "Número do Ofício"), ("date", "Data"), ("bank", "Banco"),
        ("account_base", "Conta base"), ("account_description", "Descrição conta corrente"), ("value_oficio", "Valor Ofício"),
        ("value_report", "Valor Relatório"), ("difference", "Diferença"), ("reason", "Motivo"), ("source_file", "Arquivo")
    ])
    ws = wb.create_sheet("Relatorio Bancario")
    write_sheet(ws, result["payment_report_rows"], [
        ("date", "DATA DO PAGAMENTO"), ("bank", "BANCO"), ("account_debit", "CONTA CORRENTE"), ("account_description", "DESCRIÇÃO"),
        ("cnpj_cpf", "CNPJ/CPF"), ("favorecido", "FAVORECIDO"), ("c_custo", "C.CUSTO"),
        ("conta_financeira", "CONTA FINANCEIRA"), ("valor", "VALOR")
    ])
    ws = wb.create_sheet("Transferencias")
    write_sheet(ws, result["transfer_report_rows"], [
        ("date", "DATA DA TRANSFERÊNCIA"), ("bank", "BANCO"), ("account_debit", "CONTA CORRENTE"), ("account_description", "DESCRIÇÃO"),
        ("c_custo", "CENTRO DE CUSTO"), ("conta_financeira", "CONTA FINANCEIRA"),
        ("conta_corrente_destino", "CONTA CORRENTE DESTINO"), ("valor", "VALOR")
    ])
    path = str(Path(out_dir) / f"conferencia_bancaria_{uuid.uuid4().hex}.xlsx")
    wb.save(path)
    return path


def write_sheet(ws, rows: list[dict], cols: list[tuple[str, str]]) -> None:
    fill = PatternFill("solid", fgColor="0F2D2E")
    font = Font(color="FFFFFF", bold=True)
    for col_idx, (_, title) in enumerate(cols, 1):
        c = ws.cell(1, col_idx, title)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(rows, 2):
        is_total = str(row.get("account_debit", "")).upper().startswith("TOTAL") or str(row.get("numero_oficio", "")).upper().startswith("TOTAL")
        for col_idx, (key, _) in enumerate(cols, 1):
            value = row.get(key, "")
            if key in {"value_oficio", "value_report", "difference", "valor"} and value not in (None, ""):
                value = float(value)
            c = ws.cell(row_idx, col_idx, value)
            if is_total:
                c.font = Font(bold=True)
    for col_idx in range(1, len(cols)+1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24


def process_files(files: list[InputFile]) -> dict[str, Any]:
    pdfs, sheets, ignored = split_files(files)
    registry, desc_map = load_registry(sheets)
    if not pdfs:
        raise ReconciliationError("Envie ao menos um PDF do movimento bancário.")
    oficios, report_totals, movements = [], [], []
    for f in pdfs:
        pages = pdf_pages(f)
        oficios.extend(parse_oficio_from_pdf(f, pages))
        totals, movs = parse_report_pages(f, pages)
        report_totals.extend(totals)
        movements.extend(movs)
    conciliated, divergences = [], []
    used_totals = set()
    for item in oficios:
        desc = find_description(item["account_base"], desc_map)
        base = {
            "numero_oficio": item["numero_oficio"],
            "date": item["data_br"],
            "bank": item["bank"],
            "type": item["type"],
            "account_oficio": item["account_oficio"],
            "account_base": item["account_base"],
            "account_description": desc,
            "value_oficio": item["value_oficio"],
            "source_file": item["source_file"],
        }
        if not account_in_registry(item["account_base"], registry):
            divergences.append({**base, "status": "conta_nao_cadastrada", "value_report": None, "difference": None, "reason": f"Não achei a Conta {item['account_base']} na relação de contas Bancárias."})
            continue
        match_idx = None
        match_total = None
        for idx, total in enumerate(report_totals):
            if idx in used_totals:
                continue
            if report_key_match(item, total):
                match_idx = idx
                match_total = total
                break
        if match_total is None:
            divergences.append(build_mismatch_divergence(base, item, report_totals, used_totals))
            continue
        used_totals.add(match_idx)
        vrel = float(match_total["value_report"])
        diff = round(item["value_oficio"] - vrel, 2)
        if abs(diff) <= TOL:
            conciliated.append({**base, "value_report": vrel, "difference": diff, "status": "conciliado"})
        else:
            divergences.append({**base, "status": "divergencia_valor", "value_report": vrel, "difference": diff, "reason": "Valor do ofício difere do valor totalizado no relatório."})
    conciliated = sort_conciliated(conciliated)
    divergences = sorted(divergences, key=lambda r: (_oficio_sort_key(r), r.get("date", ""), digits(r.get("account_base", "")), r.get("status", "")))

    payment_rows = [m for m in movements if m["report_type"] == "RELATÓRIO BANCÁRIO DE CONFERÊNCIA"]
    transfer_rows = [m for m in movements if m["report_type"] == "TRANSFERÊNCIA ENTRE CONTAS DA MESMA TITULARIDADE"]
    payment_rows = add_total_rows(payment_rows, "RELATÓRIO BANCÁRIO DE CONFERÊNCIA", desc_map)
    transfer_rows = add_total_rows(transfer_rows, "TRANSFERÊNCIA ENTRE CONTAS DA MESMA TITULARIDADE", desc_map)
    result = {
        "summary": {
            "pdf_count": len(pdfs), "plan_count": len(registry), "oficio_count": len(oficios), "report_total_count": len(report_totals),
            "conciliated_count": len(conciliated), "divergence_count": len(divergences),
            "oficio_total": round(sum(float(x["value_oficio"]) for x in oficios), 2),
            "report_total": round(sum(float(x["value_report"]) for x in report_totals), 2),
            "conciliated_value_oficio_total": round(sum(float(x.get("value_oficio") or 0) for x in conciliated), 2),
            "conciliated_value_report_total": round(sum(float(x.get("value_report") or 0) for x in conciliated), 2),
            "system_version": "Fase 1 - Pagamentos e Transferências v1.1",
            "last_conference_date": date_br(max([x.get("data", "") for x in oficios] or [""])),
        },
        "conciliated": conciliated,
        "divergences": divergences,
        "payment_report_rows": payment_rows,
        "transfer_report_rows": transfer_rows,
        "ignored_files": ignored,
    }
    result["excel_path"] = write_excel(result)
    return result
